import math
import pathlib

import pandas
from openpyxl import load_workbook

from .connection import Connection
from .fleet import Fleet
from .node import Node
from .path import Path


def _read_float(value: float) -> float:
    # reading excel sometimes moves integers a little (for some reason), correct this
    return round(value) if math.isclose(round(value), value, rel_tol=1e-9) else value


class MFN:
    def __init__(self, path: str | pathlib.Path):
        """
        Internal representation of a Multi Floor Network Excel Schema file.

        :param path: the path to the Excel file in MFN format
        """

        self.path = path

        self.nodes = []
        self.paths = []
        self.connections = []
        self.fleets = []

        wb = load_workbook(path, read_only=True)

        for sheet in ["NetworkNodes", "NetworkPaths", "NetworkConnections", "Fleets"]:
            if sheet not in wb.sheetnames:
                raise ValueError(f"MFN file needs a sheet called '{sheet}'")

        df = pandas.read_excel(self.path, sheet_name="NetworkNodes",
                               usecols=["name", "x_meter", "y_meter", "z_meter", "network"],
                               dtype={"name": str, "network": str},
                               header=0)
        for index, row in df.iterrows():
            self.nodes.append(
                Node(name=row["network"].lower() + "~" + row["name"].lower(), x_meter=_read_float(row["x_meter"]),
                     y_meter=_read_float(row["y_meter"]), z_meter=_read_float(row["z_meter"]),
                     network=row["network"].lower()))

        df = pandas.read_excel(self.path, sheet_name="NetworkPaths",
                               usecols=["name", "origin_node_name", "destination_node_name", "fleets", "network"],
                               dtype={"name": str, "origin_node_name": str, "destination_node_name": str, "network": str},
                               header=0)
        for index, row in df.iterrows():
            self.paths.append(
                Path(name=row["network"].lower() + "~" + row["name"].lower(),
                     origin_node_name=row["network"].lower() + "~" + row["origin_node_name"].lower(),
                     destination_node_name=row["network"].lower() + "~" + row["destination_node_name"].lower(),
                     fleets=row["fleets"], network=row["network"].lower()))

        df = pandas.read_excel(self.path, sheet_name="NetworkConnections",
                               usecols=["name", "origin_node_name", "destination_node_name",
                                        "cal_trans_duration_seconds", "fleets", "origin_network",
                                        "destination_network"],
                               dtype={"name": str, "origin_node_name": str, "destination_node_name": str, "origin_network": str, "destination_network": str},
                               header=0)
        for index, row in df.iterrows():
            self.connections.append(
                Connection(name=row["name"],
                           origin_node_name=row["origin_network"].lower() + "~" + row["origin_node_name"].lower(),
                           destination_node_name=row["destination_network"].lower() + "~" + row[
                               "destination_node_name"].lower(),
                           cal_trans_duration_seconds=int(row["cal_trans_duration_seconds"]), fleets=row["fleets"],
                           origin_network=row["origin_network"].lower(),
                           destination_network=row["destination_network"].lower()))

        df = pandas.read_excel(self.path, sheet_name="Fleets",
                               usecols=["fleet", "avg_speed_mps"],
                               header=0)
        for index, row in df.iterrows():
            self.fleets.append(Fleet(name=row["fleet"], avg_speed_mps=row["avg_speed_mps"]))
